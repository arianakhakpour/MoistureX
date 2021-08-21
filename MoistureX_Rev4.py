"""
THIS CODE ESTIMATES THE 
RELATIVE HUMIDITY WITHIN
THE LED CHAMBER BASED ON
THE CLIMATE DATA FILE 
AND OPERATING SCHEDULE.

ALL UNITS IN S.I.
BY: ARIANA KHAKPOUR
"""

#Importing the required python functions. Make sure your python is installed with
#math and openpyxl modules as well. Or you can install free version of Anacnda from
#this website: https://www.anaconda.com/distribution/
import openpyxl
import math as mth
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import math as mat


#The following lines obtains "vent area", and "climate file name" and "sheet name" from
#the user, in order for the software to find the file to access its data.
print("PLEASE NOTE THAT THIS CODE IS ONLY VALID FOR VE8 SERIES VENTS DUE TO SIMILAR THICKNESS")
print("DROPLET CONDENSATION ON THE VENTS OR CONTAMINATION AFFECTS THE PERMEABILITY PROPERTIES\nOF THE VENTS AND MAKES THIS FORECAST A LOT LESS ACCURATE")
print("___________________________________________________________________________")
print("PLEASE ENTER THE VENT'S ACTIVE AREA IN SQUARE MILIMETER \n(NOTE: THE CURRENT LED MODEL USES GORETEX VENT VE8-0308 WITH AN ACTIVE VENTING AREA OF 8.55 MM^2) \n\nAREA = ")
area = float(input())
print("\nPLEASE NOTE THAT THE FILE NAME IS CASE SENSITIVE. FAILURE TO ENTER THE \nEXACT FILE NAME WILL RESULT IN AN EXECUTION ERROR. \nALSO NOTE THAT THERE MUST BE ONLY ONE FILE WITH THIS NAME\nON YOUR SYSTEM.\n")
print("*(A)* PLEASE ENTER THE EXACT FILE NAME FOR THE CLIMATE DATA FILE FOLLOWED BY .xlsx = ")
filenom = str(input())
print("\nLoading File... This may take a while...\n")
cfd = openpyxl.load_workbook(filenom)

print("\n\nPLEASE NOTE THAT THE SHEET NAME IS CASE SENSITIVE. FAILURE TO ENTER THE \nEXACT SHEET NAME WILL RESULT IN AN EXECUTION ERROR. \n")
print("**(B)** PLEASE ENTER THE EXACT SHEET NAME OF THE CLIMATE EXCEL FILE ON WHICH THE HOUR \nBY HOUR DATA IS STORED = ")
sheetnom = str(input())
sheet0 = cfd.get_sheet_by_name(sheetnom)
print("___________________________________________________________________________")



#This correction factor is used to update the vapor passage based on
#vent area. For the tested vents of area 8.55mm^2, this factor is 1.
CORR = area * (0.1169590643)

#The following loop will ask the user the LED switch on and switch off time and will keep asking
#until the user confirms the input.
confirm = 'N'
while (confirm == 'N' or confirm == 'n' or confirm == 'No' or confirm == 'no'):
    valid = 0
    while (valid == 0):
        print("\nNOTE: ALL HOURS ENTERED SHALL BE IN 24H FORMAT REAL NUMBER W/DECIMALS,\nFOR EXAMPLE 7.00 REFERES TO 7:00 AM OR 15.75 REFERS TO 3:45 PM")
        print("\nENTER THE HOUR THAT THE LED IS SCHEDULED TO TURN ON: ")
        h_on = float(input())
        if (h_on < 24 and h_on >= 0):
            valid = 1
        elif (h_on >= 24):
            print("\nINVALID ENTRY, PLEASE TRY AGAIN ACCORDING TO INSTRUCTIONS")
            valid = 0
        elif (h_on < 0):
            print("\nINVALID ENTRY, PLEASE TRY AGAIN ACCORDING TO INSTRUCTIONS")
            valid = 0
        else:
            print("\nINVALID ENTRY, PLEASE TRY AGAIN ACCORDING TO INSTRUCTIONS")
            valid = 0
    
    valid = 0
    while (valid == 0):
        print("\nENTER THE HOUR THAT THE LED IS SCHEDULED TO TURN OFF: ")
        h_off = float(input())
        if (h_off < 24 and h_off >= 0):
            valid = 1
        elif (h_off >= 24):
            print("INVALID ENTRY, PLEASE TRY AGAIN ACCORDING TO INSTRUCTIONS")
            valid = 0
        elif (h_off < 0):
            print("INVALID ENTRY, PLEASE TRY AGAIN ACCORDING TO INSTRUCTIONS")
            valid = 0
        else:
            print("INVALID ENTRY, PLEASE TRY AGAIN ACCORDING TO INSTRUCTIONS")
            valid = 0
    h_on_hh = (int(h_on))
    if h_on_hh == 0:
        h_on_hh = '00'
    elif h_on_hh == 1:
        h_on_hh = '01'
    elif h_on_hh == 2:
        h_on_hh = '02'
    elif h_on_hh == 3:
        h_on_hh = '03'
    elif h_on_hh == 4:
        h_on_hh = '04'
    elif h_on_hh == 5:
        h_on_hh = '05'
    elif h_on_hh == 6:
        h_on_hh = '06'
    elif h_on_hh == 7:
        h_on_hh = '07'
    elif h_on_hh == 8:
        h_on_hh = '08'
    elif h_on_hh == 9:
        h_on_hh = '09'
    
    if (h_on >= 1):     
        h_on_mm = int((h_on % int(h_on)) * 60)
    else: 
        h_on_mm = int((h_on) * 60)
        
    if h_on_mm == 0:
        h_on_mm = '00'
    elif h_on_mm == 1:
        h_on_mm = '01'
    elif h_on_mm == 2:
        h_on_mm = '02'
    elif h_on_mm == 3:
        h_on_mm = '03'
    elif h_on_mm == 4:
        h_on_mm = '04'
    elif h_on_mm == 5:
        h_on_mm = '05'
    elif h_on_mm == 6:
        h_on_mm = '06'
    elif h_on_mm == 7:
        h_on_mm = '07'
    elif h_on_mm == 8:
        h_on_mm = '08'
    elif h_on_mm == 9:    
        h_on_mm = '09'
        
    h_off_hh = (int(h_off)) 
    if h_off_hh == 0:
        h_off_hh = '00'
    elif h_off_hh == 1:
        h_off_hh = '01'
    elif h_off_hh == 2:
        h_off_hh = '02'
    elif h_off_hh == 3:
        h_off_hh = '03'
    elif h_off_hh == 4:
        h_off_hh = '04'
    elif h_off_hh == 5:
        h_off_hh = '05'
    elif h_off_hh == 6:
        h_off_hh = '06'
    elif h_off_hh == 7:
        h_off_hh = '07'
    elif h_off_hh == 8:
        h_off_hh = '08'
    elif h_off_hh == 9:
        h_off_hh = '09'   
   
    if (h_off >= 1):     
        h_off_mm = int((h_off % int(h_off)) * 60)
    else: 
        h_off_mm = int((h_off) * 60)
        
    if h_off_mm == 0:
        h_off_mm = '00'
    elif h_off_mm == 1:
        h_off_mm = '01'
    elif h_off_mm == 2:
        h_off_mm = '02'
    elif h_off_mm == 3:
        h_off_mm = '03'
    elif h_off_mm == 4:
        h_off_mm = '04'
    elif h_off_mm == 5:
        h_off_mm = '05'
    elif h_off_mm == 6:
        h_off_mm = '06'
    elif h_off_mm == 7:
        h_off_mm = '07'
    elif h_off_mm == 8:
        h_off_mm = '08'
    elif h_off_mm == 9:  
        h_off_mm = '09'
        
    print("\n___________________________________________________________________________")        
    print("HOUR THAT LED IS TURNED ON = ", h_on, " (OR) In HH:MM:SS format = ", h_on_hh, ":", h_on_mm, ": 00")
    print("HOUR THAT LED IS TURNED OFF = ", h_off, " (OR) In HH:MM:SS format = ", h_off_hh, ":", h_off_mm, ": 00")
    print("\nDO YOU CONFIRM THIS INFO. ?")
    print("Enter Y for Yes, or N for No")
    confirm = str(input())
    print("\n***************************************************************************")        
print("***************************************************************************")
print("**********CALCULATION IN PROGRESS: THIS MAY TAKE SEVERAL MINUTES***********")
print("***************************************************************************")
print("***************************************************************************")


#The following function will approximate the water vapor concentration rate of 
#change in gr/m^3-minutes  based on the difference in vapor concentrations in the 
#two envirnoments and based on experimental results. The actual rates may "significantly" 
#vary due to air pollution and possible clogging over extented use in moist envirnoments.
def VCRPM_CALC (delta_C):
    
    d_C = abs(delta_C)  #yields the positive value (magnitude)
    VCRPS = CORR * 0.0000376342 * mat.exp(0.15292548311 * d_C) #vapor concentration change per seconds
    
    VCRPM = 60  * VCRPS  #vapor concentration change per minutes
    
    return VCRPM

#The following function is a psychometric function that estimates the vapor mass
#concentration based on the relative humidity and dry bulb temperature.
def V_CONCENTRATION_CALC (rh , t):
    a = 0.00000003230010727 * rh - 0.0000000003012727272
    b = 0.00000144926 * rh + 0.00000003169090909
    c = 0.00007509705455 * rh + 0.0000002527272727
    d = 0.003926678182 * rh - 0.00005416363636
    e = 0.05287762545 * rh + 0.00008172727273
    vc = (((a)*(t*t*t*t)) + ((b)*(t*t*t)) + ((c)*(t*t)) + ((d)*(t)) + e)
    if vc <= 0:
        vc = 0
    vc = round(vc, 5)
    return vc



#The following function is a psychometric function that estimates the relative humidity
#based on the vapor concentration and dry bulb temperature.
def RH_CALC (vc , t):
    mrx = (0.000000001102596701*(t*t*t*t)) - (0.000000476784208667*(t*t*t)) + (0.000128305099896850*(t*t)) - (0.030008109869684000*t) + (1.314954264157040000)
    mry = 10 ** mrx
    rh = mry * vc
    rh = round(rh, 3)
    return rh


#The following function aims to detect the operative status of the LEDs throught the
#entire hours in a year, and based on the user-defined switch-on and switch-off time.
#The value of returned variable OP is 1, for hours in which the LED is working and 0 when it is off.
def OPERATION_STAT(hour, h_on, h_off):
    if (h_on <= h_off):
        day_factor = int(hour/24)
        cum_hour_on = h_on + (day_factor * 24)
        cum_hour_off = h_off + (day_factor * 24)
        if (hour < cum_hour_on):
            OP = 0
        elif (hour > cum_hour_off):
            OP = 0
        else:
            OP = 1            
    if (h_on > h_off):
        day_factor = int(hour/24)
        cum_hour_on = h_on + (day_factor * 24)
        cum_hour_off = h_off + (day_factor * 24)
        if (hour < cum_hour_on and hour > cum_hour_off):
            OP = 0
        else:
            OP = 1    
    return OP



#The following function updates the internal vapor concentration for the next hour
#based on the previous internal vapor concentration, and the ambient (external) vapor
#concentration. To increase accuracy (and since the rate of vapor transfer is a function
#of difference in vapor concentrations) a minute by minute iterative approach is used, and
#the result at the end of the 60 iterations is reported as the next hour's internal vapor
#concentration.
def VC_UPDATER (int_vc, ext_vc):
    for iror in range (1, 60):  
        delta_C = int_vc - ext_vc
        VCRPM = VCRPM_CALC(delta_C)
        if (ext_vc > (int_vc)):        
            int_vc = int_vc + VCRPM
        elif (ext_vc <= (int_vc)):
            int_vc = int_vc - VCRPM
    return int_vc



#This function is used to retrieve the ambient relative humidity value for each hour from the 
#climate data file excel file.
def RH_Retriever (hour):
    i = int(hour) + 1
    rh = sheet0.cell (row = i, column = 6).value
    return rh



#This function is used to retrieve the ambient temperature value for each hour from the 
#climate data file excel file.
def Temp_Retriever (hour):
    i = int(hour) + 1
    t = sheet0.cell(row=i, column=4).value
    return t


#From here on, this makes the main body of the code. It creates a new excel file with columns named
#as (1)Hour (2)External Temp (3)External RH  (4)External Vapor Conc.  (5) Internal Temp
# (6)Internal RH  (7)Internal Vapor Conc.
    
hour = 1
ihf = openpyxl.Workbook()
ihf.create_sheet(index = 0 , title = 'LED Internal Humidity Condition Forecast')
sheet1 = sheet1 = ihf.get_sheet_by_name('LED Internal Humidity Condition Forecast')
sheet1['A1'] = 'HOUR'
sheet1['B1'] = 'EXT TEMP'
sheet1['C1'] = 'EXT RH'
sheet1['D1'] = 'EXT V.CONC'
sheet1['E1'] = 'INT TEMP'
sheet1['F1'] = 'INT RH'
sheet1['G1'] = 'INT V.CONC'

#This loop starts from hour 1 and ends at hour 8760, last hour of a year.
#the values for hour, external temp and external RH are placed in the new
#excel file and ambient vapor concentrations are calculated and placed in its columns.
for hour in range (1 , 8760):
    i = int(hour) + 1           #row number of the excel file is hour plus one. Plus one is added since the first row is filled with titles of each column.
    sheet1.cell(row=i, column=1).value = hour      #places hour number in column 1
    
    ext_temp = Temp_Retriever (hour)               #calls temp retriever function to obtain the ambient temp from the climate data file loaded earlier.
    sheet1.cell(row=i, column=2).value = ext_temp  #places the external temperature value in column 2
    
    ext_rh = RH_Retriever (hour)                   #calls relative humidity retriever function to obtain the ambient RH from the climate data file loaded earlier.
    sheet1.cell(row=i, column=3).value = ext_rh    #places the external RH value in column 3
    
    ext_vc = V_CONCENTRATION_CALC (ext_rh , ext_temp)    #calls the psychometric function defined earlier to calculate the external vapor concentrations based on the external temp and RH.
    sheet1.cell(row=i, column=4).value = ext_vc          #places the external vapor concentrations value in column 4
    
    if hour == 1:                                       #For the first hour, all internal parameters were set equal to external (ambient) parameters.
        int_temp = ext_temp                             #and the values for internal temp, RH and vapor concentrations were placed in columns 5, 6 and 7 respectively.
        sheet1.cell(row=i, column=5).value = ext_temp
        
        int_rh = ext_rh
        sheet1.cell(row=i, column=6).value = ext_rh

        int_vc = ext_vc
        sheet1.cell(row=i, column=7).value = ext_vc
    else:
        hourp = hour - 1
        OPP = OPERATION_STAT(hourp, h_on, h_off) 
        OP = OPERATION_STAT(hour, h_on, h_off)       #for hours greater than one:
        DIV = OP - OPP
        if (DIV == 1):     #detects the hours just after LED switched on to account for the VC gain due to unstable conditions
            int_vprime = VC_UPDATER(int_vc, ext_vc)    
            int_vc = int_vprime * 1.3         #Internal vapor concentration hike post rapid transient mode (turn on)
            sheet1.cell(row=i, column=7).value = int_vc   #internal vapor concentration then placed in column 7 for each corresponding row.
        elif (DIV == -1):
            int_vprime = VC_UPDATER(int_vc, ext_vc)    
            int_vc = int_vprime * 0.85        #Internal vapor concentration loss post rapid transient mode (turn off)
            sheet1.cell(row=i, column=7).value = int_vc   #internal vapor concentration then placed in column 7 for each corresponding row.
        else:
            int_vc = VC_UPDATER(int_vc, ext_vc) #Internal vapor concentration calculater function was called to calculate the internal VC at the next hour, based on external conditions.
            sheet1.cell(row=i, column=7).value = int_vc #internal vapor concentration then placed in column 7 for each corresponding row.

        if(OP == 0):                                  #After finding the internal vapor concentrations, and to detect condensation, we aim to calculate the internal relative humidity.
            int_temp = ext_temp                       #To find the RH, internal temperature must be known on top of the vapor concentration. Internal temperature is a function of external temperature and light's operative status.
        elif(OP == 1):                                #OP value for each hour is determined by calling the OPERATION_STAT function that was defined to detect hours for which the light is in operation.
            int_temp = (-0.002447543006*(ext_temp)*(ext_temp)) + (0.950598039598*(ext_temp)) + 15.474593962999  #if the ligh is in operation, its temperature can be estimated by this experimental equation. (Found in NO WIND testing)
        sheet1.cell(row=i, column=5).value = int_temp #This command places the estimated internal temp in column 5
        
        int_rh = RH_CALC (int_vc , int_temp)    #Psychometric function RH_CALC is called to calculate the internal relative humidity based on the internal vapor concentration and LED internal temperature.
        if int_rh >= 100:                       #Since this is approximation, RH values of slightly over 100% may be returned. Thus, this command reports everything above hundred as 100%.
            int_rh = 100
        elif int_rh < 0:
            int_rh = 0
        sheet1.cell(row=i, column=6).value = int_rh        #calculated RH is then placed in column 6, for the corresponding row i.
        if (int_rh >= 90):                                  #If the reported RH is over 90%, the cell is highlighted in red and the risky hour is reported.
            sheet1.cell(row=i, column=6).fill = PatternFill("solid", fgColor="FF2F44")
            print("Elevated risk of condensation detected at hour number:", hour)
        
ihf.save('Internal_Humidity_File.xlsx')            #Saving the results as the output excel file.
print("\n^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^")   
print("^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^")     
print("AFTER COMPLETION OF THIS SIMULATION, AN EXCEL FILE CONTAINING THE RESULTS\nWILL BE SAVED IN THE SAME DIRECTORY THE RAW CLIMATE DATA EXISTS. \nTHE FILE NAME WILL BE 'INTERNAL_HUMIDITY_FILE.xlsx'")
print("^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^")
print("^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^")